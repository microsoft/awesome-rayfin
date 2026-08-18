"""The availability workbook's colours must follow what the planner types.

Run: `python tools/tests/test_availability_format.py` (a script, not pytest — PLAN §19).
"""

from __future__ import annotations

import io
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / "server"))

import availability_xlsx as ax  # noqa: E402
from openpyxl import load_workbook  # noqa: E402
from schedule_store import ScheduleStore  # noqa: E402

FAILURES: list[str] = []


def check(name: str, ok: bool, detail: str = "") -> None:
    print(f"  {'ok  ' if ok else 'FAIL'}  {name}" + (f" — {detail}" if detail else ""))
    if not ok:
        FAILURES.append(name)


store = ScheduleStore.load("oth")
wb = load_workbook(io.BytesIO(ax.build_template(store)))
ws = wb[ax.SHEET]

# ── the colour is a RULE, not paint ───────────────────────────────────────────────────────
# ⚠️ THE DEFECT THIS GUARDS. The fills used to be written onto each cell at export, so they
# described the state the cell HAD when the file was generated. Change "gesperrt" to "frei" and
# you got the word `frei` on a red background — in a file whose entire purpose is to be edited by
# hand, and whose reader trusts the colour precisely because it is a colour.
ranges = list(ws.conditional_formatting)
check("the sheet carries conditional formatting at all", bool(ranges))

rules = [rule for rng in ranges for rule in rng.rules]
check("one rule per state", len(rules) >= 3, f"{len(rules)} rules")

wanted = {f'"{label}"' for label in ax.LABELS.values()}
got = {f for rule in rules for f in (rule.formula or [])}
check("each rule keys on the German label the planner sees", wanted <= got,
      f"missing {sorted(wanted - got)}")

# ⚠️ Quoted inside the formula, or Excel reads `frei` as a defined name and the rule never fires
# — a file that looks right in openpyxl and is stubbornly grey in Excel.
check("the labels are quoted for Excel", all(f.startswith('"') and f.endswith('"') for f in wanted))

# ⚠️ BOTH fg AND bg, and this is the assertion that matters most. Excel reads a DIFFERENTIAL
# format's solid fill from `bgColor`, while an ordinary cell fill uses `fgColor`. A rule with only
# `fgColor` set is present, valid and inspectable here — and renders as nothing in Excel, which is
# the only place this file is ever opened.
def _colours(rule) -> tuple[str | None, str | None]:
    fill = rule.dxf.fill if rule.dxf else None
    if fill is None:
        return None, None
    fg = fill.fgColor.rgb if fill.fgColor else None
    bg = fill.bgColor.rgb if fill.bgColor else None
    return fg, bg


pairs = [_colours(rule) for rule in rules]
check("the three rules paint three different colours",
      len({fg for fg, _ in pairs if fg}) >= 3, str(sorted(str(p) for p in pairs)))
check("every rule sets the background colour Excel actually reads",
      all(bg and bg not in ("00000000", "FF000000") for _, bg in pairs),
      str(sorted(str(bg) for _, bg in pairs)))

# ── and no cell carries a baked-in fill any more ──────────────────────────────────────────
painted = []
for row in ws.iter_rows(min_row=3, min_col=3):
    for cell in row:
        patt = cell.fill
        if patt is not None and patt.fill_type == "solid" and patt.fgColor.rgb not in (
            None, "00000000",
        ):
            painted.append(cell.coordinate)

check("no data cell has a hard-coded fill left", not painted,
      f"{len(painted)} painted, e.g. {painted[:5]}")

# The rule has to actually cover the data.
covered = " ".join(str(r.sqref) for r in ranges)
check("the rules cover the grid, starting at C3", "C3" in covered, covered[:80])

# ── and the values themselves still round-trip ────────────────────────────────────────────
# A formatting change must not quietly alter what the file says.
parsed = ax.parse_upload(store, ax.build_template(store))
check("the workbook still parses after the change",
      not parsed.get("error"), str(parsed.get("error") or ""))
check("and still reports every lecturer", len(parsed.get("teachers") or {}) == len(store.teachers),
      f"{len(parsed.get('teachers') or {})} of {len(store.teachers)}")

print()
if FAILURES:
    print(f"{len(FAILURES)} failed: " + "; ".join(FAILURES))
    raise SystemExit(1)
print("ok — the colour follows what the planner typed, not what we exported")
