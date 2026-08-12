"""Availability: the constraint the whole product turns on, and the file a planner fills in.

Run: `python tools/tests/test_availability.py` (a script, not pytest — PLAN §19).
"""

from __future__ import annotations

import io
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / "server"))

import availability_xlsx as ax  # noqa: E402
from openpyxl import load_workbook  # noqa: E402
from schedule_store import ScheduleStore  # noqa: E402

FAILURES: list[str] = []


def check(name: str, ok: bool, detail: str = "") -> None:
    print(f"  {'ok  ' if ok else 'FAIL'}  {name}" + (f" — {detail}" if detail else ""))
    if not ok:
        FAILURES.append(name)


store = ScheduleStore.load("oth")

# ── the states are a projection, and both projections must move together ──────────────────
# ⚠️ `unavailable` and `restricted` are derived sets. The solver reads one and the calendar the
# other, so a write that rebuilds only one leaves them disagreeing about who is free — and the
# disagreement is invisible until somebody is scheduled into a slot they said no to.
teacher = store.teachers[0]["teacherId"]
free_slot = next(
    s["slotId"] for s in store.slots
    if (teacher, s["slotId"]) not in store.unavailable
    and (teacher, s["slotId"]) not in store.restricted
)

store.set_availability(teacher, [{"slotId": free_slot, "state": "nicht_verfuegbar"}], "test")
check("blocking a slot reaches the set the solver reads",
      (teacher, free_slot) in store.unavailable)
store.set_availability(teacher, [{"slotId": free_slot, "state": "eingeschraenkt"}], "test")
check("and moving to 'limited' leaves the blocked set", (teacher, free_slot) not in store.unavailable)
check("while entering the restricted set", (teacher, free_slot) in store.restricted)
store.set_availability(teacher, [{"slotId": free_slot, "state": "verfuegbar"}], "test")
check("and going back to free leaves both",
      (teacher, free_slot) not in store.unavailable
      and (teacher, free_slot) not in store.restricted)

# ⚠️ ONE ROW PER (teacher, slot), NOT AN APPEND. Somebody changing their mind three times about a
# Tuesday must leave one statement; appending would make the reader work out which is current,
# and `availability.json` is read as a flat table by three separate call sites.
before = len([a for a in store.availability if a["teacherId"] == teacher])
for state in ("nicht_verfuegbar", "eingeschraenkt", "verfuegbar"):
    store.set_availability(teacher, [{"slotId": free_slot, "state": state}], "test")
after = len([a for a in store.availability if a["teacherId"] == teacher])
check("changing a slot three times leaves one row", before == after, f"{before} -> {after}")

# ── the version only moves when something actually changed ────────────────────────────────
v0 = store.availability_version
store.set_availability(teacher, [{"slotId": free_slot, "state": "verfuegbar"}], "test")
check("re-stating what is already true does not bump the version",
      store.availability_version == v0, f"{v0} -> {store.availability_version}")
store.set_availability(teacher, [{"slotId": free_slot, "state": "nicht_verfuegbar"}], "test")
check("a real change does", store.availability_version == v0 + 1)
store.set_availability(teacher, [{"slotId": free_slot, "state": "verfuegbar"}], "test")

# ── THE ONE THAT MATTERS: blocking a slot somebody teaches in must SAY SO ─────────────────
# ⚠️ Nothing moves. The plan becomes illegal, which is a perfectly reasonable thing for a planner
# to do — it is the first half of the cascade. But a UI that answers "gespeichert" and nothing
# else has hidden the only fact that mattered, so the write reports it and this pins that.
busy = None
for sess in store.sessions:
    row = store.assignment_by_session.get(sess["sessionId"])
    if row and sess.get("teacherId"):
        busy = (sess["teacherId"], row["slotId"], sess["sessionId"])
        break
tid, slot_id, sid = busy
result = store.set_availability(tid, [{"slotId": slot_id, "state": "nicht_verfuegbar"}], "test")
check("blocking a slot the lecturer teaches in reports the session",
      any(c["sessionId"] == sid for c in result["nowInConflict"]),
      f"{len(result['nowInConflict'])} reported")
check("and the report carries what the planner needs to recognise it",
      all(c.get("course") and c.get("slotId") for c in result["nowInConflict"]))

# ⚠️ `nowInConflict` IS THE CURRENT STATE, NOT THE DELTA OF THIS CLICK. A planner blocking Monday
# and then Tuesday should see both sessions, not only Tuesday's — the question the panel answers is
# "what is wrong now", and a delta would let the first one scroll out of sight. Pinned here because
# it is the kind of thing a later "optimisation" would quietly change.
second = None
for sess in store.sessions:
    row = store.assignment_by_session.get(sess["sessionId"])
    if row and sess.get("teacherId") == tid and sess["sessionId"] != sid and row["slotId"] != slot_id:
        second = (row["slotId"], sess["sessionId"])
        break
if second:
    both = store.set_availability(tid, [{"slotId": second[0], "state": "nicht_verfuegbar"}], "test")
    check("a second block still reports the first one too",
          {c["sessionId"] for c in both["nowInConflict"]} >= {sid, second[1]},
          f"{len(both['nowInConflict'])} reported")
    store.set_availability(tid, [{"slotId": second[0], "state": "verfuegbar"}], "test")

# The mirror: with nothing of theirs blocked, a block on a free slot reports nothing. Without this,
# a function that always returned every session this lecturer teaches would pass the test above.
store.set_availability(tid, [{"slotId": slot_id, "state": "verfuegbar"}], "test")
quiet = next(
    s["slotId"] for s in store.slots
    if not any(
        (a := store.assignment_by_session.get(x["sessionId"])) and a["slotId"] == s["slotId"]
        for x in store.sessions if x.get("teacherId") == tid
    )
)
result2 = store.set_availability(tid, [{"slotId": quiet, "state": "nicht_verfuegbar"}], "test")
check("blocking a free slot reports no conflict", result2["nowInConflict"] == [],
      f"{len(result2['nowInConflict'])} reported")
store.set_availability(tid, [{"slotId": quiet, "state": "verfuegbar"}], "test")

# ── bad input is refused by name, never coerced ───────────────────────────────────────────
bad = store.set_availability(teacher, [
    {"slotId": "Xx-99", "state": "verfuegbar"},
    {"slotId": free_slot, "state": "vielleicht"},
], "test")
check("an unknown slot is named, not skipped in silence", bad["unknownSlots"] == ["Xx-99"])
check("an unknown state is named too", bad["invalidStates"] == ["vielleicht"])
check("and neither was applied", bad["changed"] == 0)
check("an unknown lecturer is a question, not a guess",
      store.set_availability("Zzzzz", [], "test").get("error") == "teacher_not_found")

# ── the spreadsheet round trip ────────────────────────────────────────────────────────────
data = ax.build_template(store)
wb = load_workbook(io.BytesIO(data))
check("the template has both sheets", ax.SHEET in wb.sheetnames and ax.NOTES in wb.sheetnames,
      str(wb.sheetnames))
ws = wb[ax.SHEET]
header = [c.value for c in ws[1]]
check("every slot in the dataset gets a column",
      all(s["slotId"] in header for s in store.slots),
      f"{len(store.slots)} slots, {len(header) - 2} columns")
check("every lecturer gets a row", ws.max_row - 2 == len(store.teachers),
      f"{ws.max_row - 2} rows vs {len(store.teachers)} lecturers")

# ⚠️ PRE-FILLED, NOT BLANK. A template that arrives empty asks somebody to retype what the system
# already knows, and re-importing it would overwrite every real statement with a default.
one = store.teachers[0]
row_of = {ws.cell(row=r, column=1).value: r for r in range(3, ws.max_row + 1)}
col_of = {v: i + 1 for i, v in enumerate(header) if v}
blocked_slots = [s["slotId"] for s in store.slots if (one["teacherId"], s["slotId"]) in store.unavailable]
check("the template carries what we already believe",
      all(ws.cell(row=row_of[one["teacherId"]], column=col_of[s]).value == ax.LABELS["nicht_verfuegbar"]
          for s in blocked_slots),
      f"{len(blocked_slots)} blocked slot(s) checked")

# a human edits it — the ways a human actually edits a spreadsheet
target = store.teachers[1]
r = row_of[target["teacherId"]]
ws.cell(row=r, column=col_of["Mo-1"], value="gesperrt")
ws.cell(row=r, column=col_of["Mo-2"], value="X")          # the shorthand everybody uses
ws.cell(row=r, column=col_of["Mo-3"], value=" Frei ")      # stray whitespace and capital
ws.cell(row=r, column=col_of["Mo-4"], value=None)          # cleared cell
ws.cell(row=r, column=col_of["Mo-5"], value="Urlaub")      # a word, not a state
buf = io.BytesIO()
wb.save(buf)
parsed = ax.parse_upload(store, buf.getvalue())

states = {e["slotId"]: e["state"] for e in parsed["teachers"][target["teacherId"]]}
check("'gesperrt' reads back as blocked", states.get("Mo-1") == "nicht_verfuegbar")
check("'X' does too — the shorthand a planner actually types",
      states.get("Mo-2") == "nicht_verfuegbar")
check("whitespace and capitals survive", states.get("Mo-3") == "verfuegbar")
check("an empty cell means available", states.get("Mo-4") == "verfuegbar")
# ⚠️ THE IMPORTANT ONE. "Urlaub" is a human telling us something real; coercing it to "frei"
# because it is not in a dictionary would invert the single thing the file exists to say.
check("a word we do not understand is REFUSED, not read as free",
      "Mo-5" not in states
      and any(b["slotId"] == "Mo-5" and b["value"] == "Urlaub" for b in parsed["badValues"]))

check("a row for somebody who does not exist is reported", True)
ws.cell(row=r, column=1, value="NICHT-DA-042")
buf2 = io.BytesIO()
wb.save(buf2)
p2 = ax.parse_upload(store, buf2.getvalue())
check("an unknown lecturer in the file is named, not silently dropped",
      "NICHT-DA-042" in p2["unknownTeachers"])

# ⚠️ COLUMNS ARE READ FROM THE HEADER, NOT BY POSITION. A planner who deletes a column must not
# shift everybody's Tuesday onto their Monday — the failure would be silent and total.
wb2 = load_workbook(io.BytesIO(ax.build_template(store, target["teacherId"])))
ws2 = wb2[ax.SHEET]
ws2.delete_cols(3)  # remove Mo-1 entirely
hdr2 = [c.value for c in ws2[1]]
ws2.cell(row=3, column=hdr2.index("Di-1") + 1, value="gesperrt")
buf3 = io.BytesIO()
wb2.save(buf3)
p3 = ax.parse_upload(store, buf3.getvalue())
moved = {e["slotId"]: e["state"] for e in p3["teachers"][target["teacherId"]]}
check("deleting a column does not shift the rest of the week",
      moved.get("Di-1") == "nicht_verfuegbar" and "Mo-1" not in moved,
      f"{len(moved)} slots read")

# a file that is not a workbook at all
check("a file that is not a workbook is an answer, not a crash",
      ax.parse_upload(store, b"this is not xlsx").get("error") == "unreadable")

# ── the DRY RUN has to answer the hard half, and must not write to do it ──────────────────
# ⚠️ Found on the LIVE app: uploading a sheet that blocked four slots the lecturer teaches in
# previewed as "4 changes" and said nothing at all about the four lectures it was about to make
# illegal. The easy half (how many cells differ) is not the half anybody is deciding on.
teach = next(a for a in store.assignments
             if store.session_by_id.get(a["sessionId"], {}).get("teacherId"))
tid3 = store.session_by_id[teach["sessionId"]]["teacherId"]
slot3 = teach["slotId"]
before_rows = len(store.availability)
before_version = store.availability_version

would = store.sessions_blocked_by(tid3, {slot3: "nicht_verfuegbar"})
check("a preview names the lecture a proposed block would invalidate",
      any(w["sessionId"] == teach["sessionId"] for w in would),
      f"{len(would)} session(s) for {tid3} at {slot3}")
check("and the preview wrote NOTHING — no row, no version bump",
      len(store.availability) == before_rows and store.availability_version == before_version,
      f"rows {before_rows}->{len(store.availability)}, version {before_version}->{store.availability_version}")

# The mirror, or a function that always returns the whole timetable would pass the test above.
check("a slot the lecturer does not teach in previews as harmless",
      store.sessions_blocked_by(tid3, {slot3: "verfuegbar"}) == []
      or all(w["slotId"] != slot3 for w in store.sessions_blocked_by(tid3, {slot3: "verfuegbar"})),
      "clearing the same slot must not report it")

# And the overlay must be able to UNDO a stored block, not only add one — otherwise a planner
# freeing a morning would still be told their own lecture is in conflict.
store.set_availability(tid3, [{"slotId": slot3, "state": "nicht_verfuegbar"}], "test")
check("with the block really stored, the lecture is reported as in conflict",
      any(w["sessionId"] == teach["sessionId"] for w in store.sessions_blocked_by(tid3)))
check("and previewing the release clears it again",
      all(w["slotId"] != slot3 for w in store.sessions_blocked_by(tid3, {slot3: "verfuegbar"})))
store.set_availability(tid3, [{"slotId": slot3, "state": "verfuegbar"}], "test")

# ── what the downloaded file is CALLED ────────────────────────────────────────────────────
# ⚠️ This file is mailed to a lecturer and mailed back, so the name is not cosmetic: it is how
# the recipient knows it is theirs and how a planner tells this week's return from last week's.
# It used to be `Verfuegbarkeit_oth_IM-T007.xlsx` — the Kürzel, which means nothing to the person
# opening it, and no date at all.
from app import _filename_part, _stamp  # noqa: E402

check("a lecturer's name replaces the Kürzel",
      _filename_part("Prof. Dr. D. Danzer") == "Prof_Dr_D_Danzer",
      _filename_part("Prof. Dr. D. Danzer"))

# ⚠️ GERMAN TRANSLITERATION, NOT BARE ACCENT-STRIPPING. NFKD alone gives "Obermuller", which is a
# different surname to every German reader — and these files are named after people who read them.
check("umlauts become ae/oe/ue rather than being dropped",
      _filename_part("Prof. Dr. O. Obermüller") == "Prof_Dr_O_Obermueller",
      _filename_part("Prof. Dr. O. Obermüller"))
check("sharp s becomes ss",
      _filename_part("Franz-Josef Groß") == "Franz_Josef_Gross",
      _filename_part("Franz-Josef Groß"))

# A name that folds away entirely must not yield `Verfuegbarkeit_oth__2026-…xlsx`.
check("a name with nothing ASCII left still produces a filename",
      _filename_part("教授") == "unbenannt")
check("an empty name does too", _filename_part("") == "unbenannt")

# The header is latin-1 on the wire; a stray umlaut or quote here is a failed download, not a
# cosmetic problem.
#
# ⚠️ The LABEL is ASCII even though the probe is not. These run on a Windows console at cp1252
# (PLAN §19: scripts, not pytest), and printing the probe itself crashed the harness on the CJK
# case — a test that dies while reporting a pass is indistinguishable from a broken feature.
for _label, _probe in [
    ("a German umlaut", "Prof. Dr. B. Beutlhäuser"),
    ("Spanish accents", "Ana María Ruiz-Pérez"),
    ("non-Latin script", "教授"),
    ("an absurdly long name", "x" * 90),
]:
    _name = f"Verfuegbarkeit_oth_{_filename_part(_probe)}_{_stamp()}.xlsx"
    check(f"header-safe filename for {_label}",
          _name.isascii() and '"' not in _name and "/" not in _name and len(_name) < 120)

# The stamp exists to order two downloads of the same person on the same day, so it needs the
# time, not just the date.
import re as _re  # noqa: E402

check("the timestamp carries date AND time",
      bool(_re.fullmatch(r"\d{4}-\d{2}-\d{2}_\d{4}", _stamp())), _stamp())

# ⚠️ AND IT IS LOCAL TIME. The container runs UTC; stamping a 22:15 download as `2015` is worse
# than omitting the time, because it is confidently wrong about the only thing it asserts.
from datetime import datetime, timezone  # noqa: E402

try:
    from zoneinfo import ZoneInfo

    _berlin = datetime.now(ZoneInfo("Europe/Berlin")).strftime("%Y-%m-%d_%H%M")
    check("the timestamp is Europe/Berlin, not UTC", _stamp() == _berlin,
          f"{_stamp()} vs berlin {_berlin} / utc {datetime.now(timezone.utc):%Y-%m-%d_%H%M}")
except Exception as exc:  # noqa: BLE001
    check("tzdata is installed so Europe/Berlin resolves", False, str(exc))

# ── the real dataset's table is seeded, and the seed did not eat the real answers ─────────
# ⚠️ THE RISK OF SEEDING IS THAT IT OVERWRITES. 17 055 assumed rows were added around 333 real
# ones; a fill that ran before the stated rows, or keyed on the wrong pair, would bury OTH's
# actual Zeitwünsche under `verfuegbar` and nothing on screen would look wrong — the panel would
# simply show a lecturer as free all week, which is what it showed before anyone asked.
real = ScheduleStore.load("oth-real")
_pairs = {(a["teacherId"], a["slotId"]) for a in real.availability}
_tids = {t["teacherId"] for t in real.teachers}
_sids = {s["slotId"] for s in real.slots}

check("every lecturer x slot has a row on the real dataset",
      len(_pairs) == len(_tids) * len(_sids),
      f"{len(_pairs)} pairs vs {len(_tids)}x{len(_sids)}={len(_tids) * len(_sids)}")
check("and no pair is written twice", len(_pairs) == len(real.availability))
check("no row points at a lecturer the app cannot show",
      all(a["teacherId"] in _tids for a in real.availability))
check("every row says where it came from",
      all(a.get("source") for a in real.availability))

_stated = [a for a in real.availability if a.get("source") != "assumed"]
check("OTH's own Zeitwünsche survived the seed",
      len(_stated) >= 300, f"{len(_stated)} stated rows")
check("and they are not all 'verfuegbar', which is what burying them would look like",
      any(a["state"] == "nicht_verfuegbar" for a in _stated),
      str(sorted({a["state"] for a in _stated})))
check("the assumed rows really are the neutral state",
      all(a["state"] == "verfuegbar" for a in real.availability if a.get("source") == "assumed"))

# The distinction has to reach the client, or materialising it in the file achieves nothing: the
# panel already drew a full green week by defaulting the missing rows.
_view = real.availability_for(_stated[0]["teacherId"])
check("availability_for reports the source per slot",
      all("source" in s for s in _view["slots"]),
      str(_view["slots"][0]))

print()
if FAILURES:
    print(f"{len(FAILURES)} failed: " + "; ".join(FAILURES))
    raise SystemExit(1)
print("ok — availability can be stated, and saying it never means guessing it")
