"""Read a Stundenverteilungsplan Excel — the artefact OTH's process actually runs on.

PLAN §25.3: the Stundenverteilungsplan is assembled from **one Excel file per professor**, filled
in, collected and merged by hand, and only then typed into Untis. So this file is the UPSTREAM half
of the process, and it carries the one thing the Untis export does not: what a lecturer said about
their own availability, in their own words, before anyone scheduled anything.

Two things are read:

  * the **availability grid** (`1_Vorlesungen` J5:N11) — Mo-Fr × 6 blocks, values 1 / 0 / -1 with
    OTH's own legend printed beside it: *1 = gut möglich · 0 = wenns sein muss · -1 = NICHT
    möglich*. Three states, which is exactly the vocabulary this app already uses.
  * the **course demand** (`1_Vorlesungen` from row 17) — module, alias, total SWS across all
    groups, and the number of groups. The group count is why timetabling is hard, and it is stated
    here rather than inferred.

⚠️ THE FILE DOES NOT SAY WHICH UNTIS LECTURER IT BELONGS TO. The filename carries a human name
and Untis carries a short code; nothing in either states the link.
Guessing it from the surname is the `find_teacher` mistake with worse consequences — an
availability constraint attached to the wrong professor would move somebody else's lectures.

So the link is made from EVIDENCE: the sheet's **Alias** column holds the same subject keys Untis
uses after the faculty prefix (`BW` -> `BW_BW`, `DEN` -> `BW_DEN`, `RSM` -> `BW_RSM`), so the
lecturer is the one whose taught subjects match the aliases. A unique strong match resolves; a tie
or a weak match returns nothing and says why.
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError:  # pragma: no cover - reported, never fatal
    openpyxl = None  # type: ignore[assignment]

DAYS = ["Mo", "Di", "Mi", "Do", "Fr"]

# OTH's own legend, printed in the sheet at J13:L13.
STVP_STATE = {1: "verfuegbar", 0: "eingeschraenkt", -1: "nicht_verfuegbar"}


def _grid(ws: Any, first_row: int, first_col: int) -> list[tuple[str, int, Any]]:
    """The 5x6 availability block: returns (day, block, raw value)."""
    out = []
    for di, day in enumerate(DAYS):
        for b in range(6):
            v = ws.cell(first_row + b, first_col + di).value
            if v is not None:
                out.append((day, b + 1, v))
    return out


def read(path: Path) -> dict[str, Any]:
    """Parse one StVP workbook into availability + course demand."""
    if openpyxl is None:
        raise SystemExit("openpyxl is required to read the Stundenverteilungsplan")
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["1_Vorlesungen"] if "1_Vorlesungen" in wb.sheetnames else wb.worksheets[0]

    # ⚠️ The grid is LOCATED, not hard-coded to J6. The header row carries the weekday names, so
    # the block is found by looking for them — a template that grows a column stays readable.
    anchor = None
    for r in range(1, min(ws.max_row, 30) + 1):
        for c in range(1, min(ws.max_column, 20) + 1):
            if str(ws.cell(r, c).value or "").strip().lower() == "montag":
                if all(str(ws.cell(r, c + i).value or "").strip().lower() == DAY_NAMES[i]
                       for i in range(5)):
                    anchor = (r + 1, c)
                    break
        if anchor:
            break
    if not anchor:
        raise SystemExit(f"{path.name}: could not find the Mo-Fr availability header")

    availability = []
    unknown: list[Any] = []
    for day, block, raw in _grid(ws, *anchor):
        try:
            val = int(raw)
        except (TypeError, ValueError):
            unknown.append(raw)
            continue
        state = STVP_STATE.get(val)
        if state is None:
            unknown.append(raw)
            continue
        availability.append({"slotId": f"{day}-{block}", "state": state})

    # Course demand: find the header row by its first column label.
    courses = []
    head = None
    for r in range(1, min(ws.max_row, 40) + 1):
        if str(ws.cell(r, 2).value or "").strip().lower().startswith("modulname"):
            head = r
            break
    if head:
        for r in range(head + 1, ws.max_row + 1):
            title = ws.cell(r, 2).value
            if not title:
                continue
            courses.append({
                "title": str(title).strip(),
                "alias": str(ws.cell(r, 3).value or "").strip() or None,
                "swsTotal": ws.cell(r, 4).value,
                "groups": ws.cell(r, 5).value,
                "programmes": str(ws.cell(r, 6).value or "").strip() or None,
                "status": str(ws.cell(r, 8).value or "").strip() or None,
                "mode": str(ws.cell(r, 9).value or "").strip() or None,
            })

    person = re.sub(r"^Winter\d+-StVP-", "", path.stem).strip()
    return {
        "file": path.name,
        "person": person,
        "availability": availability,
        "courses": courses,
        "unknownValues": unknown,
    }


DAY_NAMES = ["montag", "dienstag", "mittwoch", "donnerstag", "freitag"]


def match_teacher(stvp: dict[str, Any], subjects_by_teacher: dict[str, set[str]]) -> dict[str, Any]:
    """Which Untis short code does this workbook belong to?

    ⚠️ MATCHED ON WHAT THEY TEACH, NOT ON THEIR NAME. The aliases in the sheet are the subject keys
    Untis uses after the faculty prefix, so a lecturer who teaches those subjects is the author.
    A name-based guess would be a guess; this is a join.

    Returns the code and the evidence, or `None` with a reason — never a best effort.
    """
    aliases = {a["alias"].strip().upper() for a in stvp["courses"] if a.get("alias")}
    if not aliases:
        return {"teacherId": None, "reason": "the sheet lists no subject aliases to match on"}

    scores: list[tuple[int, str, set[str]]] = []
    for teacher, subjects in subjects_by_teacher.items():
        suffixes = {s.split("_", 1)[1].upper() if "_" in s else s.upper() for s in subjects}
        hit = aliases & suffixes
        if hit:
            scores.append((len(hit), teacher, hit))
    if not scores:
        return {"teacherId": None, "reason": f"no lecturer teaches any of {sorted(aliases)}"}

    scores.sort(reverse=True)
    best = scores[0]
    rivals = [s for s in scores if s[0] == best[0]]
    if len(rivals) > 1:
        return {"teacherId": None,
                "reason": f"{len(rivals)} lecturers match {best[0]} alias(es) equally: "
                          f"{sorted(r[1] for r in rivals)}"}
    # ⚠️ One alias in common is a coincidence waiting to happen — two subjects is a person.
    if best[0] < 2:
        return {"teacherId": None,
                "reason": f"only one alias matched ({sorted(best[2])}), which is too weak to bind "
                          f"an availability constraint to {best[1]}"}
    return {"teacherId": best[1], "matchedAliases": sorted(best[2]), "aliasesInSheet": sorted(aliases)}
