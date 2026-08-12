"""The availability spreadsheet — a planner's file, not a database export.

⚠️ THE FILE IS THE INTERFACE, SO IT IS WRITTEN FOR THE PERSON FILLING IT IN. OTH's real process
(PLAN §25) is one Excel per professor, collected by mail and merged by hand — so the shape here is
deliberately the shape they already use: a person, their week, and a word in each cell. The dataset
values (`verfuegbar` / `eingeschraenkt` / `nicht_verfuegbar`) are identifiers, not prose, and never
appear in a cell.

⚠️ THE IMPORT IS FORGIVING, THE VALIDATION IS NOT. A returned spreadsheet has been touched by a
human: an "X", a "nein", a stray capital, a renamed sheet. Anything a reasonable person would mean
is accepted; anything ambiguous is REFUSED BY NAME rather than guessed at, because the cost of
guessing here is a lecturer marked free on a morning they told us they cannot teach.
"""

from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

SHEET = "Verfügbarkeit"
NOTES = "Hinweise"

#: What the planner sees ↔ what the dataset stores.
LABELS = {
    "verfuegbar": "frei",
    "eingeschraenkt": "eingeschränkt",
    "nicht_verfuegbar": "gesperrt",
}
#: Everything we will accept back. Keys are compared lower-cased and stripped.
ACCEPTED = {
    "frei": "verfuegbar", "verfügbar": "verfuegbar", "verfuegbar": "verfuegbar",
    "ja": "verfuegbar", "j": "verfuegbar", "y": "verfuegbar", "ok": "verfuegbar", "": "verfuegbar",
    "eingeschränkt": "eingeschraenkt", "eingeschraenkt": "eingeschraenkt",
    "e": "eingeschraenkt", "nur im notfall": "eingeschraenkt", "(e)": "eingeschraenkt",
    "gesperrt": "nicht_verfuegbar", "nicht verfügbar": "nicht_verfuegbar",
    "nicht_verfuegbar": "nicht_verfuegbar", "nein": "nicht_verfuegbar", "n": "nicht_verfuegbar",
    "x": "nicht_verfuegbar", "-": "nicht_verfuegbar", "keine zeit": "nicht_verfuegbar",
}

_HEAD = PatternFill("solid", fgColor="1C1917")
_HEAD_FONT = Font(color="FFFFFF", bold=True)

#: ⚠️ THESE ARE APPLIED AS CONDITIONAL FORMATTING, NOT AS CELL FILLS.
#:
#: They used to be written onto each cell once, at export. The colour then described the state the
#: cell HAD when the file was generated, and stopped agreeing with it the moment anybody typed:
#: change "gesperrt" to "frei" and you get the word `frei` on a red background. In a file whose
#: entire job is to be edited by hand, a colour that lies about the cell it sits in is worse than
#: no colour — the planner reading the returned sheet trusts the colour, because that is what
#: colour is for.
#:
#: As a rule keyed on the cell's own text, the fill follows what the person typed, in their Excel,
#: with no macro and no re-export.
#: ⚠️ BOTH COLOURS ARE SET, AND THAT IS NOT REDUNDANT. A normal cell fill draws a solid pattern
#: from `fgColor`; a DIFFERENTIAL format — which is what a conditional-formatting rule carries — is
#: read by Excel from `bgColor`. Set only one and the rule is present, valid, inspectable in
#: openpyxl, and renders as nothing at all in the application the file exists for. `start_color`
#: and `end_color` set fg and bg respectively, so the colour survives whichever one is consulted.
_FILL = {
    "verfuegbar": PatternFill(start_color="FFE7F6EC", end_color="FFE7F6EC", fill_type="solid"),
    "eingeschraenkt": PatternFill(start_color="FFFDF3D7", end_color="FFFDF3D7", fill_type="solid"),
    "nicht_verfuegbar": PatternFill(start_color="FFFBE3E3", end_color="FFFBE3E3", fill_type="solid"),
}
#: Readable text to go with each fill, so the three states differ by more than hue — the grid has
#: to be usable printed in grey and by anyone who cannot separate red from green.
_FONT = {
    "verfuegbar": Font(color="166534"),
    "eingeschraenkt": Font(color="92400E"),
    "nicht_verfuegbar": Font(color="991B1B"),
}
_THIN = Side(style="thin", color="D6D3D1")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _slot_columns(store) -> list[dict]:
    """The week, in the order the dataset defines it — never a hard-coded Mo–Fr × 7."""
    return [
        {
            "slotId": s["slotId"],
            "day": s.get("day") or s["slotId"].split("-")[0],
            "block": s.get("block"),
            "start": s.get("startTime", ""),
            "end": s.get("endTime", ""),
        }
        for s in store.slots
    ]


def build_template(store, teacher: str | None = None) -> bytes:
    """One sheet, one row per lecturer, one column per slot.

    `teacher` narrows it to a single person — the per-professor file OTH already sends round.
    """
    cols = _slot_columns(store)
    if teacher:
        t = store.find_teacher(teacher)
        teachers = [t] if t else []
    else:
        teachers = sorted(store.teachers, key=lambda x: x.get("name", ""))

    state_by = {}
    for a in store.availability:
        state_by[(a["teacherId"], a["slotId"])] = a["state"]

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET

    ws.cell(row=1, column=1, value="Kürzel").fill = _HEAD
    ws.cell(row=1, column=1).font = _HEAD_FONT
    ws.cell(row=2, column=1, value="").fill = _HEAD
    ws.cell(row=1, column=2, value="Lehrperson").fill = _HEAD
    ws.cell(row=1, column=2).font = _HEAD_FONT
    ws.cell(row=2, column=2, value="").fill = _HEAD

    # Two header rows: the day/block above, the clock time below. A planner reads the time; the
    # importer reads the slot id, which is written into the first row as the cell comment-free
    # value `Mo-1`. Both are needed and neither is guessable from the other.
    for i, c in enumerate(cols):
        col = 3 + i
        h1 = ws.cell(row=1, column=col, value=c["slotId"])
        h1.fill = _HEAD
        h1.font = _HEAD_FONT
        h1.alignment = Alignment(horizontal="center")
        h2 = ws.cell(row=2, column=col, value=f"{c['day']} {c['start']}")
        h2.fill = _HEAD
        h2.font = Font(color="FFFFFF", size=9)
        h2.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col)].width = 11

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 34
    ws.freeze_panes = "C3"

    dv = DataValidation(
        type="list",
        formula1='"' + ",".join(LABELS.values()) + '"',
        allow_blank=True,
        showDropDown=False,
    )
    dv.error = "Bitte frei, eingeschränkt oder gesperrt wählen."
    dv.errorTitle = "Unbekannter Wert"
    ws.add_data_validation(dv)

    for r, t in enumerate(teachers, start=3):
        ws.cell(row=r, column=1, value=t["teacherId"]).font = Font(size=9, color="78716C")
        ws.cell(row=r, column=2, value=t.get("name", ""))
        for i, c in enumerate(cols):
            state = state_by.get((t["teacherId"], c["slotId"]), "verfuegbar")
            cell = ws.cell(row=r, column=3 + i, value=LABELS[state])
            cell.border = _BORDER
            cell.alignment = Alignment(horizontal="center")
        dv.add(f"C{r}:{get_column_letter(2 + len(cols))}{r}")

    # ── the colours, as rules rather than as paint ────────────────────────────────────────
    #
    # ⚠️ ONE RANGE, THREE RULES, KEYED ON THE CELL'S OWN TEXT. Excel re-evaluates these on every
    # edit, so a planner who changes "frei" to "gesperrt" sees the cell turn red as they type.
    # Writing the fill per cell instead — which is what this did — froze the colour at export time
    # and left `frei` sitting on a red background the moment anyone corrected anything.
    #
    # The comparison is against the LABEL, because the label is what is in the cell: the dataset
    # identifiers never appear in the sheet (see the module note). `"frei"` has to be quoted inside
    # the formula string or Excel reads it as a defined name and the rule silently never fires.
    if teachers:
        grid = f"C3:{get_column_letter(2 + len(cols))}{2 + len(teachers)}"
        for state, label in LABELS.items():
            ws.conditional_formatting.add(
                grid,
                CellIsRule(
                    operator="equal",
                    formula=[f'"{label}"'],
                    fill=_FILL[state],
                    font=_FONT[state],
                ),
            )

    note = wb.create_sheet(NOTES)
    for i, line in enumerate(
        [
            f"Verfügbarkeiten — {store.site_label}",
            "",
            "So füllen Sie die Datei aus:",
            "  • Pro Zeile eine Lehrperson, pro Spalte ein Zeitfenster.",
            "  • Erlaubte Werte: frei · eingeschränkt · gesperrt",
            "    (akzeptiert werden beim Hochladen auch: x, nein, ja, e, leer)",
            "  • „eingeschränkt“ heißt: möglich, aber nur wenn es nicht anders geht.",
            "",
            "Bitte NICHT ändern:",
            "  • die Spalte „Kürzel“ (Zeile 1) — daran wird die Lehrperson erkannt",
            "  • die erste Kopfzeile mit den Zeitfenster-Kürzeln (Mo-1, Mo-2 …)",
            "  • den Namen des Tabellenblatts",
            "",
            "Zeilen dürfen gelöscht werden: wer nicht in der Datei steht, bleibt unverändert.",
            "",
            "Die Datei ändert nur, WANN jemand unterrichten kann — sie verschiebt keine",
            "Veranstaltungen. Wenn ein Zeitfenster gesperrt wird, in dem bereits unterrichtet",
            "wird, meldet die Anwendung das nach dem Hochladen und schlägt eine Umplanung vor.",
        ],
        start=1,
    ):
        note.cell(row=i, column=1, value=line)
    note.column_dimensions["A"].width = 92

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_upload(store, data: bytes) -> dict[str, Any]:
    """Read a returned spreadsheet into `{teacherId: [{slotId, state}]}` plus everything wrong with it.

    ⚠️ NOTHING IS APPLIED HERE. Parsing and writing are separate acts so the caller can show the
    planner what a file would do before it does it — the same reason `/api/rules/preview` exists in
    PLAN §26 and the same reason a proposal is not a publish.
    """
    try:
        wb = load_workbook(io.BytesIO(data), data_only=True)
    except Exception as exc:  # noqa: BLE001 — any openpyxl failure is one answer to the caller
        return {"error": "unreadable", "detail": str(exc)[:200]}

    ws = wb[SHEET] if SHEET in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        return {"error": "no_rows", "detail": f"{len(rows)} Zeilen"}

    header = rows[0]
    # Column index → slotId, taken from the header rather than assumed by position: a planner who
    # deletes a column must not silently shift everybody's Tuesday onto their Monday.
    slot_cols: dict[int, str] = {}
    unknown_cols: list[str] = []
    for i, h in enumerate(header):
        if i < 2 or h in (None, ""):
            continue
        sid = str(h).strip()
        if sid in store.slot_by_id:
            slot_cols[i] = sid
        else:
            unknown_cols.append(sid)
    if not slot_cols:
        return {"error": "no_slot_columns",
                "detail": "Die Kopfzeile enthält keine bekannten Zeitfenster-Kürzel."}

    per_teacher: dict[str, list[dict]] = {}
    names: dict[str, str] = {}
    unknown_teachers: list[str] = []
    bad_values: list[dict] = []
    for r in rows[2:]:
        if not r or all(v in (None, "") for v in r):
            continue
        key = (str(r[0]).strip() if len(r) > 0 and r[0] not in (None, "") else "") or \
              (str(r[1]).strip() if len(r) > 1 and r[1] not in (None, "") else "")
        if not key:
            continue
        t = store.find_teacher(key)
        if not t:
            unknown_teachers.append(key)
            continue
        entries: list[dict] = []
        for i, sid in slot_cols.items():
            raw = r[i] if i < len(r) else None
            text = "" if raw is None else str(raw).strip()
            state = ACCEPTED.get(text.lower())
            if state is None:
                # ⚠️ Refused by name, not coerced. "Urlaub" in a cell is a human telling us
                # something real; turning it into "frei" because it is not in a dictionary would
                # invert the one thing the file exists to say.
                bad_values.append({"teacher": t.get("name"), "slotId": sid, "value": text[:40]})
                continue
            entries.append({"slotId": sid, "state": state})
        per_teacher[t["teacherId"]] = entries
        names[t["teacherId"]] = t.get("name", "")

    return {
        "sheet": ws.title,
        "teachers": per_teacher,
        "names": names,
        "slotColumns": len(slot_cols),
        "unknownColumns": unknown_cols,
        "unknownTeachers": unknown_teachers,
        "badValues": bad_values,
    }
